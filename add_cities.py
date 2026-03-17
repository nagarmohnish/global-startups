"""Merge scraped data and add tabs to global_startups_final.xlsx."""
import json, re, sys, os, glob
sys.stdout.reconfigure(encoding="utf-8")
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

COL_ORDER = ["Name","Website","Industry","Description","Founded","Funding","Last Round","Founders","Top Investors","Team Size"]

def normalize_name(name):
    n = name.lower().strip()
    n = re.sub(r"\s*\(.*?\)", "", n)
    n = re.sub(r"\s*(ltd\.?|inc\.?|co\.?|corp\.?|pte\.?|ag|gmbh|sa|sas|sl|llc|ab|corporation|technologies|tech|security|ai)\s*\.?$", "", n, flags=re.IGNORECASE)
    return re.sub(r"[^a-z0-9]", "", n)

def merge_deduplicate(sources_list):
    seen = {}
    for source in sources_list:
        for row in source:
            key = normalize_name(row.get("Name",""))
            if not key:
                continue
            if key in seen:
                existing = seen[key]
                for col in COL_ORDER:
                    if (not existing.get(col) or existing[col] in ("N/A","")) and row.get(col) and row[col] not in ("N/A",""):
                        existing[col] = row[col]
            else:
                seen[key] = {col: row.get(col,"") for col in COL_ORDER}
                seen[key]["Name"] = row.get("Name","")
    return list(seen.values())

def format_sheet(ws, df):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_fill
    col_widths = {
        "Name": 25, "Website": 35, "Industry": 15, "Description": 55,
        "Founded": 12, "Funding": 14, "Last Round": 18, "Founders": 40,
        "Top Investors": 50, "Team Size": 25,
    }
    for col_num, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_num)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ALL cities with their data file patterns (rebuild all tabs)
CITIES_TO_ADD = [
    # Existing tabs (rebuild with updated data)
    ("Tel Aviv", ["f6s_telaviv"]),
    ("Beijing", ["seedtable_beijing"]),
    ("Seoul", ["seedtable_seoul"]),
    ("Singapore", ["failory_singapore", "seedtable_singapore", "f6s_singapore"]),
    ("Shanghai", ["failory_shanghai", "seedtable_shanghai"]),
    ("Tokyo", ["seedtable_tokyo", "f6s_tokyo"]),
    ("Paris", ["failory_paris", "seedtable_paris"]),
    ("Zurich", ["failory_zurich", "seedtable_zurich", "f6s_zurich"]),
    ("Berlin", ["failory_berlin", "seedtable_berlin", "f6s_berlin"]),
    ("Sao Paulo", ["failory_saopaulo", "seedtable_saopaulo", "f6s_saopaulo"]),
    ("Madrid", ["failory_madrid", "seedtable_madrid", "f6s_madrid"]),
    ("Silicon Valley", ["failory_sf_siliconvalley", "failory_sj_siliconvalley", "seedtable_sf_siliconvalley", "seedtable_sj_siliconvalley"]),
    # New tabs
    ("Shenzhen", ["failory_shenzhen", "seedtable_shenzhen", "f6s_china_shenzhen"]),
    ("NYC", ["failory_new_york_city", "seedtable_new_york", "f6s_united-states_new-york"]),
    ("London", ["failory_london", "seedtable_london", "f6s_united-kingdom_london"]),
    ("Boston", ["failory_boston", "seedtable_boston", "f6s_united-states_boston"]),
    ("Los Angeles", ["failory_los_angeles", "seedtable_los_angeles", "f6s_united-states_los-angeles"]),
    ("Hangzhou", ["failory_hangzhou", "seedtable_hangzhou", "f6s_china_hangzhou"]),
    ("Stockholm", ["failory_stockholm", "seedtable_stockholm", "f6s_sweden_stockholm"]),
    ("Guangzhou", ["failory_guangzhou", "seedtable_guangzhou", "f6s_china_guangzhou"]),
]

# Load workbook
for candidate in ["global_startups_final.xlsx", "global_startups_final_v2.xlsx"]:
    if os.path.exists(candidate):
        src = candidate
        break
wb = load_workbook(src)
print(f"Loaded: {src} with sheets: {wb.sheetnames}")

total_added = 0
for tab_name, patterns in CITIES_TO_ADD:
    sources = []
    for pattern in patterns:
        fname = f"data/{pattern}.json"
        if os.path.exists(fname):
            with open(fname, encoding="utf-8") as f:
                data = json.load(f)
                if data:
                    sources.append(data)
                    print(f"  {fname}: {len(data)}")

    if not sources:
        print(f"  {tab_name}: no data files found, skipping")
        continue

    merged = merge_deduplicate(sources)
    df = pd.DataFrame(merged)
    df = df.drop_duplicates(subset=["Name"], keep="first")
    for col in COL_ORDER[1:]:
        if col in df.columns:
            df[col] = df[col].replace("", "N/A")
    df["Description"] = df["Description"].apply(lambda x: (x[:300]+"...") if isinstance(x,str) and len(x)>300 else x)
    df = df[[c for c in COL_ORDER if c in df.columns]]

    sheet = tab_name[:31]
    if sheet in wb.sheetnames:
        del wb[sheet]

    ws = wb.create_sheet(sheet)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    format_sheet(ws, df)
    print(f"  {tab_name}: {len(df)} startups added")
    total_added += len(df)

outfile = "global_startups_final.xlsx"
try:
    wb.save(outfile)
except PermissionError:
    outfile = "global_startups_final_v2.xlsx"
    wb.save(outfile)
    print(f"(saved as {outfile} - original locked)")

print(f"\nAll sheets: {wb.sheetnames}")
print(f"New cities added: {total_added} startups")
