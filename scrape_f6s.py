"""
Scrape F6S Tel Aviv startups listing and export to Excel.
Uses Playwright with stealth to bypass bot detection,
then extracts structured data from .company-block DOM elements.
Fetches real website URLs from F6S profiles (batch via new tabs)
and falls back to Google search for bot-blocked ones.
"""

import re
import time
import pandas as pd
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


SOCIAL_DOMAINS = {
    "f6s.com", "f6s.ca", "linkedin.com", "facebook.com", "twitter.com",
    "x.com", "hubspot.com", "instagram.com", "youtube.com", "github.com",
    "crunchbase.com", "angel.co", "wellfound.com", "google.com",
}


def create_browser(p):
    """Create a stealth browser context."""
    browser = p.chromium.launch(
        headless=False,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
    )
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/131.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1920, "height": 1080},
        locale="en-US",
    )
    page = context.new_page()
    stealth = Stealth()
    stealth.apply_stealth_sync(page)
    return browser, context, page


def scrape_startups():
    """Scrape startup data from F6S Tel Aviv page using DOM selectors."""
    url = "https://www.f6s.com/companies/israel/tel-aviv/lo"

    with sync_playwright() as p:
        browser, context, page = create_browser(p)

        print("Loading page...")
        for attempt in range(3):
            try:
                page.goto(url, timeout=30000)
                break
            except Exception as e:
                print(f"  Attempt {attempt+1} failed: {e}")
                if attempt == 2:
                    raise
                page.wait_for_timeout(3000)
        page.wait_for_timeout(5000)

        # Scroll to load all companies
        print("Scrolling to load all companies...")
        for _ in range(10):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1000)

        print("Extracting startup data from DOM...")
        blocks = page.query_selector_all(".company-block")
        print(f"Found {len(blocks)} company blocks")

        startups = []
        for block in blocks:
            startup = extract_company(block)
            if startup["Name"]:
                startups.append(startup)

        # Phase 1: Extract domains from company names
        needs_lookup = []
        for startup in startups:
            domain = extract_domain_from_name(startup["Name"])
            if domain:
                startup["Website"] = f"https://{domain}"
            else:
                needs_lookup.append(startup)

        print(f"\n{len(startups) - len(needs_lookup)} URLs extracted from names")
        print(f"{len(needs_lookup)} need profile lookup")

        # Phase 2: Batch fetch F6S profiles using multiple tabs
        bot_blocked = fetch_websites_from_profiles(context, needs_lookup)

        # Phase 3: Google search for remaining companies
        if bot_blocked:
            print(f"\n{len(bot_blocked)} companies bot-blocked, using Google search...")
            fetch_websites_from_google(page, bot_blocked)

        browser.close()
        return startups


def fetch_websites_from_profiles(context, startups):
    """Open F6S profiles in batches of tabs to get real website URLs."""
    bot_blocked = []
    batch_size = 5

    for batch_start in range(0, len(startups), batch_size):
        batch = startups[batch_start:batch_start + batch_size]
        tabs = []

        # Open tabs in parallel
        for startup in batch:
            f6s_url = startup.get("F6S URL", "")
            if not f6s_url:
                continue
            tab = context.new_page()
            try:
                tab.goto(f6s_url, timeout=15000, wait_until="domcontentloaded")
            except Exception:
                tab.close()
                bot_blocked.append(startup)
                continue
            tabs.append((tab, startup))

        # Wait for pages to load
        time.sleep(3)

        # Extract websites from each tab
        for tab, startup in tabs:
            try:
                body_text = tab.inner_text("body")
                if "bot" in body_text.lower() and len(body_text) < 500:
                    bot_blocked.append(startup)
                    print(f"  [blocked] {startup['Name']}")
                else:
                    website = get_website_from_profile(tab)
                    if website:
                        startup["Website"] = website
                        print(f"  [profile] {startup['Name']} -> {website}")
                    else:
                        bot_blocked.append(startup)
                        print(f"  [no url]  {startup['Name']}")
            except Exception:
                bot_blocked.append(startup)
            finally:
                tab.close()

    return bot_blocked


def fetch_websites_from_google(page, startups):
    """Use Google search to find real website URLs for companies."""
    for i, startup in enumerate(startups):
        name = startup["Name"]
        # Clean name for search
        clean_name = re.sub(r"\s*\(.*?\)", "", name).strip()
        query = f"{clean_name} startup Tel Aviv official website"

        try:
            search_url = f"https://www.google.com/search?q={query.replace(' ', '+')}"
            page.goto(search_url, timeout=15000)
            page.wait_for_timeout(2000)

            # Extract the first non-social result URL
            links = page.query_selector_all("a[href]")
            for link in links:
                href = link.get_attribute("href") or ""
                if (
                    href.startswith("http")
                    and not is_social_or_internal(href)
                    and not "google" in href
                    and not "cache" in href
                    and not "webcache" in href
                ):
                    startup["Website"] = href
                    print(f"  [google]  {name} -> {href}")
                    break
            else:
                startup["Website"] = ""
                print(f"  [google]  {name} -> not found")
        except Exception as e:
            startup["Website"] = ""
            print(f"  [google]  {name} -> error: {e}")


def is_social_or_internal(href):
    """Check if a URL belongs to a social/internal domain."""
    for domain in SOCIAL_DOMAINS:
        if domain in href:
            return True
    return False


def extract_domain_from_name(name):
    """Extract domain from company name like 'Wonderful (wonderful.ai)' -> 'wonderful.ai'."""
    match = re.search(r"\(([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\)", name)
    if match:
        return match.group(1)
    return None


def get_website_from_profile(page):
    """Extract the company's real website URL from their F6S profile page."""
    links = page.query_selector_all("a[href]")
    for link in links:
        href = link.get_attribute("href") or ""
        if href and href.startswith("http") and not is_social_or_internal(href):
            return href
    return ""


def extract_company(block):
    """Extract structured data from a single .company-block element."""
    startup = {
        "Name": "",
        "Tagline": "",
        "Description": "",
        "Location": "",
        "Founded": "",
        "Funding": "",
        "Investors": "",
        "Team Members": "",
        "Website": "",
        "F6S URL": "",
    }

    # Name
    name_el = block.query_selector("h2.company-entry-title a")
    if name_el:
        startup["Name"] = name_el.inner_text().strip()
        startup["F6S URL"] = name_el.get_attribute("href") or ""

    # Tagline
    tagline_el = block.query_selector("h3")
    if tagline_el:
        startup["Tagline"] = tagline_el.inner_text().strip()

    # Description
    desc_el = block.query_selector(".profile-description")
    if desc_el:
        startup["Description"] = desc_el.inner_text().strip()
        if startup["Description"].endswith("\nmore"):
            startup["Description"] = startup["Description"][:-5].strip()

    # Location
    location_text = block.evaluate(
        """el => {
            const divs = el.querySelectorAll('.centered-content');
            for (const div of divs) {
                const svg = div.querySelector('use[*|href="#location"]');
                if (svg) return div.textContent.trim();
            }
            return '';
        }"""
    )
    startup["Location"] = location_text.strip()

    # Founded
    founded_text = block.evaluate(
        """el => {
            const divs = el.querySelectorAll('.centered-content');
            for (const div of divs) {
                const svg = div.querySelector('use[*|href="#clock"]');
                if (svg) return div.textContent.trim();
            }
            return '';
        }"""
    )
    founded_match = re.search(r"Founded\s+(\d{4})", founded_text)
    if founded_match:
        startup["Founded"] = founded_match.group(1)

    # Funding & Investors
    funding_text = block.evaluate(
        """el => {
            const divs = el.querySelectorAll('.centered-content');
            for (const div of divs) {
                const svg = div.querySelector('use[*|href="#trend"]');
                if (svg) return div.textContent.trim();
            }
            return '';
        }"""
    )
    if funding_text:
        funding_match = re.search(r"(\$[\d,.]+[kmb]?)", funding_text, re.IGNORECASE)
        if funding_match:
            startup["Funding"] = funding_match.group(1)

        investor_match = re.search(
            r"(?:Raised\s*from|raised\s*from)\s*(.*?)(?:See all|$)",
            funding_text,
            re.IGNORECASE,
        )
        if investor_match:
            investors = investor_match.group(1).strip()
            investors = re.sub(r"\s*and\s*\d+\s*more\s*$", "", investors).strip()
            startup["Investors"] = investors

    # Team members
    team_el = block.query_selector(".collection-team-summary-wrapper")
    if team_el:
        team_text = team_el.inner_text().strip()
        meet_match = re.search(r"Meet\s+(.*?)\s+that\s+work", team_text)
        if meet_match:
            startup["Team Members"] = meet_match.group(1)

    return startup


def create_excel(startups, filename="f6s_tel_aviv_startups.xlsx"):
    """Create a formatted Excel file from the startup data."""
    df = pd.DataFrame(startups)

    # Remove duplicates
    df = df.drop_duplicates(subset=["Name"], keep="first")

    # Replace empty strings with N/A
    for col in ["Founded", "Funding", "Investors", "Team Members", "Description", "Website"]:
        df[col] = df[col].replace("", "N/A")

    # Truncate long descriptions
    df["Description"] = df["Description"].apply(
        lambda x: (x[:300] + "...") if isinstance(x, str) and len(x) > 300 else x
    )

    # Reorder columns
    col_order = [
        "Name",
        "Website",
        "Tagline",
        "Description",
        "Location",
        "Founded",
        "Funding",
        "Investors",
        "Team Members",
        "F6S URL",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    # Write to Excel with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Startups")
        ws = writer.sheets["Startups"]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = alt_fill

        col_widths = {
            "Name": 35, "Website": 35, "Tagline": 45, "Description": 60,
            "Location": 20, "Founded": 12, "Funding": 15, "Investors": 40,
            "Team Members": 25, "F6S URL": 50,
        }
        for col_num, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_num)
            ws.column_dimensions[letter].width = col_widths.get(col_name, 20)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"\nExcel file saved: {filename}")
    print(f"Total startups: {len(df)}")
    return df


if __name__ == "__main__":
    print("=" * 60)
    print("F6S Tel Aviv Startups Scraper")
    print("=" * 60)

    startups = scrape_startups()

    print(f"\nParsed {len(startups)} startups")

    if startups:
        for s in startups[:3]:
            print(f"  - {s['Name']}: {s['Website']}")

        df = create_excel(startups)
        print("\nFirst 15 entries:")
        print(df[["Name", "Website", "Founded"]].head(15).to_string(index=False))
    else:
        print("No startups found. The page structure may have changed.")
