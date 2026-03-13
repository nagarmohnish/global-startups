"""
Scrape Failory Tel Aviv startups listing and export to Excel.
Extracts data from both the featured startup detail sections
and the main table, combining into a single clean dataset.
"""

import sys
import re
import json
import pandas as pd
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# Known website mappings for the 10 featured startups (from Failory's own links)
FEATURED_WEBSITES = {
    "Flytrex": "https://flytrex.com",
    "Earnix": "https://earnix.com",
    "Appcharge": "https://appcharge.com",
    "Red Access": "https://redaccess.io",
    "Noma Security": "https://noma.security",
    "Qodo": "https://qodo.ai",
    "Remedio": "https://remedio.io",
    "Firefly": "https://firefly.ai",
    "Aidoc": "https://aidoc.com",
    "Firebolt": "https://firebolt.io",
}


def scrape_failory():
    """Scrape startup data from Failory Tel Aviv page."""
    url = "https://www.failory.com/startups/tel-aviv"

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
        )
        page = context.new_page()
        stealth = Stealth()
        stealth.apply_stealth_sync(page)

        print("Loading Failory page...")
        page.goto(url, timeout=30000)
        page.wait_for_timeout(5000)

        # Scroll to load everything
        for _ in range(15):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1000)

        # === Extract featured startups (top 10 with detail tables) ===
        print("Extracting featured startups...")
        featured = extract_featured(page)

        # === Extract table startups (90 rows) ===
        print("Extracting table startups...")
        table_startups = extract_table(page)

        browser.close()

        # Combine, deduplicate
        all_names = set()
        combined = []
        for s in featured:
            if s["Name"] and s["Name"] not in all_names:
                all_names.add(s["Name"])
                combined.append(s)
        for s in table_startups:
            if s["Name"] and s["Name"] not in all_names:
                all_names.add(s["Name"])
                combined.append(s)

        print(f"Total: {len(combined)} startups ({len(featured)} featured + {len(table_startups)} table)")
        return combined


def extract_featured(page):
    """Extract the 10 featured startups with detail tables."""
    # Get numbered headings
    headings = page.query_selector_all("h3")
    featured_names = []
    for h in headings:
        text = h.inner_text().strip()
        match = re.match(r"(\d+)\.\s+(.+)", text)
        if match:
            featured_names.append(match.group(2).strip())

    # Get detail tables
    detail_tables = page.query_selector_all("table:not(.failory-table)")
    featured = []

    for i, table in enumerate(detail_tables):
        rows = table.query_selector_all("tr")
        data = {}
        for row in rows:
            cells = row.query_selector_all("td, th")
            if len(cells) >= 2:
                key = cells[0].inner_text().strip()
                val = cells[1].inner_text().strip()
                data[key] = val

        # Get description from paragraph before the table
        desc = table.evaluate("""el => {
            let prev = el.previousElementSibling;
            while (prev) {
                if (prev.tagName === 'P' && prev.textContent.trim().length > 20) {
                    return prev.textContent.trim();
                }
                prev = prev.previousElementSibling;
            }
            return '';
        }""")

        name = featured_names[i] if i < len(featured_names) else ""
        website = FEATURED_WEBSITES.get(name, "")

        startup = {
            "Name": name,
            "Website": website,
            "Description": desc[:500] if desc else "",
            "Industry": "",
            "Founded": data.get("Year Founded", ""),
            "Funding": data.get("Funding Amount", ""),
            "Last Round": data.get("Last Funding Status", ""),
            "Team Size": data.get("Startup Size", ""),
            "Founders": data.get("Founders", ""),
            "Top Investors": data.get("Top Investors", ""),
        }
        featured.append(startup)
        print(f"  {name}: {website}")

    return featured


def extract_table(page):
    """Extract startups from the main failory-table."""
    rows = page.query_selector_all("table.failory-table tbody tr")
    startups = []

    for row in rows:
        name_el = row.query_selector("a.startup-name")
        industry_el = row.query_selector("td.industry")
        year_el = row.query_selector("td.year")
        amount_el = row.query_selector("td.amount")
        round_el = row.query_selector("td.round")

        name = name_el.inner_text().strip() if name_el else ""
        href = name_el.get_attribute("href") if name_el else ""
        if href and "?ref=" in href:
            href = href.split("?ref=")[0]

        startup = {
            "Name": name,
            "Website": href or "",
            "Description": "",
            "Industry": industry_el.inner_text().strip() if industry_el else "",
            "Founded": year_el.inner_text().strip() if year_el else "",
            "Funding": amount_el.inner_text().strip() if amount_el else "",
            "Last Round": round_el.inner_text().strip() if round_el else "",
            "Team Size": "",
            "Founders": "",
            "Top Investors": "",
        }
        if name:
            startups.append(startup)

    return startups


def create_excel(startups, filename="tel_aviv_startups.xlsx"):
    """Create a formatted Excel file from the startup data."""
    df = pd.DataFrame(startups)

    # Remove duplicates
    df = df.drop_duplicates(subset=["Name"], keep="first")

    # Replace empty strings with N/A
    na_cols = ["Website", "Description", "Industry", "Founded", "Funding",
               "Last Round", "Team Size", "Founders", "Top Investors"]
    for col in na_cols:
        if col in df.columns:
            df[col] = df[col].replace("", "N/A")

    # Truncate long descriptions
    df["Description"] = df["Description"].apply(
        lambda x: (x[:300] + "...") if isinstance(x, str) and len(x) > 300 else x
    )

    # Reorder columns
    col_order = [
        "Name", "Website", "Industry", "Description", "Founded",
        "Funding", "Last Round", "Founders", "Top Investors", "Team Size",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    # Write to Excel with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tel Aviv Startups")
        ws = writer.sheets["Tel Aviv Startups"]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = alt_fill

        col_widths = {
            "Name": 25, "Website": 35, "Industry": 15, "Description": 55,
            "Founded": 12, "Funding": 14, "Last Round": 18, "Founders": 40,
            "Top Investors": 50, "Team Size": 25,
        }
        for col_num, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_num)
            ws.column_dimensions[letter].width = col_widths.get(col_name, 20)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"\nExcel file saved: {filename}")
    print(f"Total startups: {len(df)}")
    return df


if __name__ == "__main__":
    print("=" * 60)
    print("Failory Tel Aviv Startups Scraper")
    print("=" * 60)

    startups = scrape_failory()

    if startups:
        df = create_excel(startups)
        print("\nFirst 15 entries:")
        print(df[["Name", "Website", "Industry", "Funding"]].head(15).to_string(index=False))
    else:
        print("No startups found.")
